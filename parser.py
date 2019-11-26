import pandas as pd
import requests
import re
import os
import base64
import json
from time import sleep
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.styles import PatternFill
from itertools import islice
from bs4 import BeautifulSoup
from requests.exceptions import ReadTimeout
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from docx_edit import create_letter_from_template

SOURCE_EXCEL_BOOK = 'input-data.xlsx'

# columns which we expect to find in source excel document
COLUMNS = {
    'fio': 'фио должника',
    'bank': "банк!!!!",
    'inn': "инн",
    'ln': "кд"
}

# websites which are the information source
COURTS_DECISIONS_SOURCE = 'http://www.reyestr.court.gov.ua'
COURTS_INFO_SOURCE = 'https://court.gov.ua/sudova-vlada/sudy/'
COURTS_FULL_INFO_REQUEST_URL = 'https://court.gov.ua/search_court.php'

# pattern for search court's full address
ADDRESS_MARK = "Адреса:"

# chrome driver path
CHROME_DRV_PATH = "chrome.drv/chromedriver.exe"

# constant part in the name of results directory
RES_DIR_PART_NAME = "RESULTS-"


def get_real_columns_names(columns):
    real_names = {}
    key_list = list(COLUMNS.keys())
    val_list = list(COLUMNS.values())
    for col in columns:
        curr_col = str(col).lower().strip()
        if curr_col in COLUMNS.values():
            c = {
                key_list[val_list.index(curr_col)]: col
            }
            real_names.update(c)
    # print("realNAMES!!!!", real_names)
    return real_names


def parse_source_excel_doc():
    wb = load_workbook(filename=SOURCE_EXCEL_BOOK)
    data = wb.active.values

    max_row = wb.active.max_row
    # wb.close()

    columns_real_names = {}

    for row_ in range(1, max_row + 1):
        cols = next(data)
        # print(cols)
        if any(cols):
            cols_low = [str(col).lower().strip() for col in cols]
            # print('low',cols_low)
            # print([True if defined_col in cols_low else False for defined_col in COLUMNS.values()])
            # print(all([True if defined_col in cols_low else False for defined_col in COLUMNS.values()]))
            if all([True if defined_col in cols_low else False for defined_col in COLUMNS.values()]):
                print(get_real_columns_names(cols))
                columns_real_names.update(get_real_columns_names(cols))
                break
    # print("cols", cols)
    if not columns_real_names:
        print('ERROR: Source document is not valid')
        exit()

    data = list(data)
    data = (islice(r, 0, None) for r in data)
    df = pd.DataFrame(data, columns=cols)

    wb.close()

    df['decision-number'] = 'NONE'
    df['decision-link'] = 'NONE'
    df['casenumber'] = 'NONE'
    df['courtname'] = 'NONE'
    df['regdate'] = 'NONE'
    df['errors'] = 'NONE'
    return {'dataframe': df, 'columns': columns_real_names}


def save_results_book(dataframe, start_time):
    wb = load_workbook(filename=SOURCE_EXCEL_BOOK)
    results_sheet = wb.create_sheet("Results")
    results_sheet.insert_rows(1, amount=5)
    results_sheet.cell(row=1, column=1, value='Started at:')
    results_sheet.cell(row=1, column=2, value=start_time)
    results_sheet.cell(row=1, column=3, value='Finished at:')
    results_sheet.cell(row=1, column=4, value=datetime.now())

    for r in dataframe_to_rows(dataframe, index=True, header=True):
        results_sheet.append(r)

    errors_col = ''
    for row in results_sheet.iter_rows(min_row=1, max_col=results_sheet.max_column, max_row=results_sheet.max_row):
        for cell in row:
            if cell.value == 'errors':
                errors_col = cell
                break
        if errors_col:
            print("BREAK!!!!")
            break

    print("COLLLLL", errors_col.column)
    for col in results_sheet.iter_cols(min_col=errors_col.column, max_col=errors_col.column, min_row=1,
                                       max_row=results_sheet.max_row, values_only=False):
        for cell in col:
            if cell.value not in ("NONE", None):
                if re.search('ERROR', cell.value):
                    print(cell.value)
                    cell.fill = PatternFill("solid", fgColor="DE1010")
                elif re.search('WARNI', cell.value):
                    print(cell.value)
                    cell.fill = PatternFill("solid", fgColor="FF9800")

    wb.save(SOURCE_EXCEL_BOOK)
    wb.close()


def get_web_source_status():
    d_session = requests.Session()
    c_session = requests.Session()
    try:
        d_r = d_session.get(COURTS_DECISIONS_SOURCE, timeout=(10, 10))
        if not 200 <= d_r.status_code < 300:
            print(f'host {COURTS_DECISIONS_SOURCE} have a problem')
            return False
    except ReadTimeout:
        print(f'host {COURTS_DECISIONS_SOURCE} is unavailable')
        return False

    try:
        c_r = c_session.get(COURTS_INFO_SOURCE, timeout=(10, 10))
        if not 200 <= c_r.status_code < 300:
            print(f'host {COURTS_INFO_SOURCE} have a problem')
            return False
    except ReadTimeout:
        print(f'host {COURTS_INFO_SOURCE} is unavailable')
        return False

    return True


# ConnectionError

# ---------------------------------------------- WEB---------------------------------------
def get_court_address(court_name):
    payload = {
        'q_court_search': court_name,
        'type_page': 'main_page'
    }

    session = requests.Session()

    try:
        courts_info_request = session.post(COURTS_FULL_INFO_REQUEST_URL, data=payload)
        if not 200 <= courts_info_request.status_code < 300:
            print(f'host {COURTS_FULL_INFO_REQUEST_URL} have a problem')
            return f'ERRROR!!!! host {COURTS_FULL_INFO_REQUEST_URL} have a problem'

        courts_info_soup = BeautifulSoup(courts_info_request.text, 'html.parser')
        specific_court_url = courts_info_soup.find(name='a')['href']
        courts_full_info_request = session.get(specific_court_url)

        if not 200 <= courts_full_info_request.status_code < 300:
            print(f'host {COURTS_FULL_INFO_REQUEST_URL} have a problem')
            return f'ERRROR!!!! host {COURTS_FULL_INFO_REQUEST_URL} have a problem'
    except ReadTimeout:
        print(f'host  {COURTS_FULL_INFO_REQUEST_URL} is unavailable')
        return f'ERRROR!!!! host {COURTS_FULL_INFO_REQUEST_URL} have a problem'

    courts_full_info_soup = BeautifulSoup(courts_full_info_request.text, 'html.parser')

    for linebreak in courts_full_info_soup.find_all('br'):
        linebreak.extract()

    paragraph_list = courts_full_info_soup.find_all('p')

    complete_address = ""

    for paragraph in paragraph_list:
        if paragraph.find(string=re.compile(ADDRESS_MARK)):
            complete_address = re.sub(ADDRESS_MARK, '', paragraph.text)
            break

    return complete_address if complete_address else "ERRORR!!!! ADDRESS not found"


def get_courts_decisions(browser_driver=None, loan_agreement=None):
    if not loan_agreement:
        return {'error': {'message': f'ERROR!!!! loan number is not defined '}}
    browser_driver.get(COURTS_DECISIONS_SOURCE)
    search_input = browser_driver.find_element_by_id('SearchExpression')
    button_input = browser_driver.find_element_by_id('btn')

    if not search_input or not button_input:
        return {'error': {'message': f'ERROR!!!! Site {COURTS_DECISIONS_SOURCE} is unavailable'}}

    search_input.send_keys(loan_agreement)
    button_input.submit()

    decisions_soup = BeautifulSoup(browser_driver.page_source, 'html.parser')
    if not decisions_soup.title.text == "Єдиний державний реєстр судових рішень":
        return {'error': {'message': f'ERROR!!!! Site {COURTS_DECISIONS_SOURCE} is unavailable'}}

    results_table = decisions_soup.find(attrs={"id": "divresult"})
    if not results_table:
        return {'error': {'message': f'WARNING!!!! Decision for loan agreement {loan_agreement}  is absent'}}

    decisions = {}
    for decision_type_tag in results_table.find_all('td', string=re.compile('Рішення')):
        # for decision_type_tag in results_table.find_all('td', string=re.compile('Ухвала')):

        parent_tag = decision_type_tag.parent
        decision_number = re.sub('/', '-', parent_tag.a.text)
        decision_number = re.sub(r"\\", '-', decision_number)
        href = parent_tag.a['href']
        regdate = parent_tag.find('td', attrs={'class': 'RegDate'}).text
        casenumber = parent_tag.find('', attrs={'class': 'CaseNumber'}).text
        court_name = parent_tag.find('', attrs={'class': 'CourtName'}).text
        decisions.update({href:
            {
                'decision-number': decision_number,
                'decision-link': href,
                'casenumber': re.sub("\n", '', casenumber).strip(),
                'courtname': re.sub("\n", '', court_name).strip(),
                'regdate': re.sub("\n", '', regdate).strip()
            }
        })

    if len(decisions) == 0:
        decisions.update({'error': {'message': f'WARNING!!!! Decision for loan agreement {loan_agreement}  is absent'}})
    print("DEcision", decisions)
    return decisions


def create_headless_chrome(chromedriver=CHROME_DRV_PATH, headless=False):
    webdriver_options = Options()

    if headless:
        webdriver_options.add_argument('--headless')
        webdriver_options.add_argument('--disable-gpu')

    driver = webdriver.Chrome(chromedriver, options=webdriver_options)

    # driver = webdriver.Chrome('chrome.drv/chromedriver.exe')

    return driver


def fill_decisions_dataframe(source_dataframe=None, frame_columns=None, headless_browser=None):
    loan_column_name = frame_columns['ln']
    # 1. copy structure of existent dataframe

    result_dataframe = source_dataframe.iloc[0:0]
    # 2. iterate over rows in the source dataframe
    for index, row in source_dataframe.iterrows():
        single_row = source_dataframe.iloc[index:index + 1].reset_index(drop=True)

        court_decisions_search_result = get_courts_decisions(headless_browser, row[loan_column_name])

        if court_decisions_search_result.get('error', None):
            single_row.loc[0, 'errors'] = court_decisions_search_result['error']['message']
            result_dataframe = result_dataframe.append(single_row, ignore_index=True)
        else:
            # 3. copy appropriate row from source dataframe in exact quantity  how many decisions were found
            for value_ in court_decisions_search_result.values():
                single_row.loc[0, 'regdate'] = value_['regdate']
                single_row.loc[0, 'decision-link'] = value_['decision-link']
                single_row.loc[0, 'decision-number'] = value_['decision-number']
                single_row.loc[0, 'courtname'] = value_['courtname']
                single_row.loc[0, 'casenumber'] = value_['casenumber']
                result_dataframe = result_dataframe.append(single_row, ignore_index=True)

    return result_dataframe


def create_results_dir():
    curr_dir = os.getcwd()

    try:
        index = 0
        while True and index < 60:
            time_now = datetime.now()
            results_dir_name = curr_dir + '\\' + \
                               RES_DIR_PART_NAME + re.sub('-', '_', str(time_now.date())) + \
                               '-' + re.sub(':', '_', str(str(time_now.time())))[:8]
            if not os.path.exists(results_dir_name):
                os.mkdir(results_dir_name)
                break
            else:
                print("Results dir exist. Sleep 1 sec")
                index += 1
                sleep(1)
        else:
            raise Exception('attempt index is full')
    except Exception as e:
        print("Results dir creation error")
        return None

    return results_dir_name


def create_dir(dir_path, dir_name):
    full_path = dir_path + '\\' + dir_name
    index = 0
    while index < 10:

        try:
            os.mkdir(full_path)
            return full_path
        except FileExistsError as e:
            print("Results dir exist. Sleep 1 sec")
            full_path += '_new'
            index += 1
            sleep(1)
        except Exception as e:
            return None

    return None


# prepare data for generate letter in docx doc
def get_template_context(row, columns):
    context_dict = {

        'fio': row[columns['fio']],
        'decision_number': row['decision-number'],
        'casenumber': row['casenumber'],
        'courtname': row['courtname'],
        'regdate': row['regdate']

    }

    return context_dict


def court_documents_generator(results_dir_path=None, results_dataframe=None, frame_columns=None, headless_browser=None):
    if results_dataframe is None or results_dir_path is None:
        return False

    for index, row in results_dataframe.iterrows():
        if row['decision-number'] not in ("NONE", None):
            borrower_dir_name = row[frame_columns['fio']] + '-' + str(int(row[frame_columns['inn']]))
            if not os.path.exists(results_dir_path + '\\' + borrower_dir_name):
                borrower_dir_path = create_dir(results_dir_path, borrower_dir_name)
            else:
                borrower_dir_path = results_dir_path + '\\' + borrower_dir_name
            if borrower_dir_path is not None:
                decision_dir_path = create_dir(borrower_dir_path, row['decision-number'])
                if decision_dir_path is not None:
                    if not create_letter_from_template(get_template_context(row, frame_columns), decision_dir_path):
                        results_dataframe.loc[index, 'errors'] += "; ERROR!!!! Letter file creation error"
                        continue

                    if not get_pdf_from_html(url_path=COURTS_DECISIONS_SOURCE + row['decision-link'],
                                             web_driver=headless_browser,
                                             print_options={},
                                             save_path=decision_dir_path,
                                             save_file_name=row['decision-number'] + '.pdf'):
                        results_dataframe.loc[index, 'errors'] += "; ERROR!!!! Decision pdf file creation error"

                        continue
                else:
                    results_dataframe.loc[index, 'errors'] += "; ERROR!!!! Borrower/(Decision) dir creation error"
                    continue


            else:
                results_dataframe.loc[index, 'errors'] += "; ERROR!!!! Borrower dir creation error"
                continue

    return results_dataframe


def get_pdf_from_html(url_path=None, web_driver=None, print_options={}, save_path="", save_file_name=""):
    def send_devtools(w_driver, cmd, params={}):
        resource = "/session/%s/chromium/send_command_and_get_result" % w_driver.session_id
        url = w_driver.command_executor._url + resource
        body = json.dumps({'cmd': cmd, 'params': params})
        response = w_driver.command_executor._request('POST', url, body)
        print(response)
        # if response['status']:
        #   raise Exception(response.get('value'))
        # return response.get('value')
        if (response.get('value') is not None):
            return response.get('value')
        else:
            return None

    def validate_file_name(save_path, save_file_name):
        index = 0
        file_name = save_file_name
        while index < 10:
            file_path = os.path.join(save_path, file_name)
            if not os.path.exists(file_path):
                return file_path
            else:
                print("File with given name exist. Generating new.")
                file_name = file_name.split('.pdf')[0] + '_new' + '.pdf'
                index += 1

        return False

    def save_as_pdf(result_data, save_path, save_file_name):
        # https://timvdlippe.github.io/devtools-protocol/tot/Page#method-printToPDF
        file_path = validate_file_name(save_path, save_file_name)
        if not file_path:
            return False
        if result_data is not None:
            with open(file_path, 'wb') as file:
                # ответ приходит в base64 - декодируем
                file.write(base64.b64decode(result_data['data']))
            return True
        else:
            return False

    if not all([url_path, web_driver, save_path, save_file_name]):
        return False

    web_driver.get(url_path)
    btn_print = web_driver.find_element_by_id('btnPrint')
    btn_print.click()

    # file_full_path = os.path.join(os.path.normpath(save_path), save_file_name)

    # задаем параметры печати
    calculated_print_options = {
        'landscape': False,
        'displayHeaderFooter': False,
        'printBackground': True,
        'preferCSSPageSize': True,
    }
    calculated_print_options.update(print_options)

    # запускаем печать в pdf файл
    result = send_devtools(web_driver, "Page.printToPDF", calculated_print_options)
    if save_as_pdf(result, save_path, save_file_name):
        return True
    else:
        return False


def main():
    prepared_data = parse_source_excel_doc()
    data_frame = prepared_data['dataframe']
    verified_columns = prepared_data['columns']
    start_time = datetime.now()
    curr_dir = os.getcwd()

    # {
    #     'fio': 'фио должника',
    #     'bank': "банк!!!!",
    #     'inn': "инн"
    # }
    # 'decision-link' = 'NONE'
    # 'casenumber' = 'NONE'
    # 'courtname' = 'NONE'
    # 'regdate' = 'NONE'
    # 'errors' = 'NONE'
    result_dir_name = RES_DIR_PART_NAME + re.sub('-', '_', str(start_time.date())) + \
                      '-' + re.sub(':', '_', str(start_time.time()))[:8]
    results_dir_path = create_dir(curr_dir, result_dir_name)

    if results_dir_path is None:
        print("System error")
        exit()

    if get_web_source_status():
        web_dr = create_headless_chrome()
        final_dataframe = fill_decisions_dataframe(source_dataframe=data_frame,
                                                   frame_columns=verified_columns,
                                                   headless_browser=web_dr)
        web_dr.close()

        web_dr = create_headless_chrome(headless=True)
        court_documents_generator(results_dir_path=results_dir_path,
                                  results_dataframe=final_dataframe,
                                  frame_columns=verified_columns,
                                  headless_browser=web_dr)

        print(final_dataframe)
        save_results_book(final_dataframe, start_time)
        web_dr.close()


main()
